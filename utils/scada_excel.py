from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
import numbers
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from utils.energy import normalize_gas_to_kwh

DEFAULT_PCI_FACTOR = 9.082
DEFAULT_INTERVAL_MINUTES = 10
DEFAULT_SHEET_NAME = "BILAN TOTAL"

DATE_ROW = 10
TIME_ROW = 11
START_ROW = 12
DATA_START_COL = 4

PCI_PATTERN = re.compile(r"([0-9]+(?:[.,][0-9]+)?)")

FIELD_ORDER: list[str | None] = [
    "gas_volume_nm3_cumul",
    None,
    "gas_flow_nm3h",
    "auxiliary_energy_kwh",
    "electrical_power_kw",
    "runtime_hours",
    "electrical_energy_kwh",
    "reactive_energy_kvarh",
    "rpm",
    "power_factor",
    "voltage_v",
    "current_phase1_a",
    "current_phase2_a",
    "current_phase3_a",
    "chilled_water_flow_m3h",
    "chilled_water_temp_in",
    "chilled_water_temp_out",
    "chilled_water_energy_kwh",
    "chilled_water_power_kw",
    "hot_water_recup_flow_m3h",
    "hot_water_recup_temp_in",
    "hot_water_recup_temp_out",
    "hot_water_recup_energy_kwh",
    "hot_water_recup_power_kw",
    "hot_water_alpha_san_flow_m3h",
    "hot_water_alpha_san_temp_in",
    "hot_water_alpha_san_temp_out",
    "hot_water_alpha_san_energy_kwh",
    "hot_water_alpha_san_power_kw",
    "hot_water_alpha_san_setpoint",
    "hot_water_alpha_san_temp_reading",
    "hot_water_alpha_flow_m3h",
    "hot_water_alpha_temp_in",
    "hot_water_alpha_temp_out",
    "hot_water_alpha_energy_kwh",
    "hot_water_alpha_power_kw",
    "hot_water_alpha_setpoint",
    "hot_water_alpha_temp_reading",
    "hot_water_gamma_flow_m3h",
    "hot_water_gamma_temp_in",
    "hot_water_gamma_temp_out",
    "hot_water_gamma_energy_kwh",
    "hot_water_gamma_power_kw",
    "hot_water_gamma_setpoint",
    "hot_water_gamma_temp_reading",
    "efficiency_electrical_pct",
    "efficiency_thermal_pct",
    "efficiency_total_pct",
    "steg_purchase_kwh",
    "steg_sale_kwh",
    "production_positive_kwh",
    "production_negative_kwh",
]


@dataclass(frozen=True)
class ScadaReading:
    timestamp: datetime
    normalized_kwh: float
    power_gross_kw: float
    gas_flow_nm3h: float
    raw_metrics: dict[str, float | None]


@dataclass(frozen=True)
class ScadaParseStats:
    total_columns: int
    parsed_columns: int
    skipped_columns: int
    pci_factor: float
    interval_minutes: int
    sheet_name: str


def parse_numeric_value(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, numbers.Number):
        return float(value)

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None

    return None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, numbers.Number):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(cleaned).date()
        except ValueError:
            return None

    return None


def _parse_time(value: Any) -> time | None:
    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, time):
        return value

    if isinstance(value, numbers.Number):
        try:
            return from_excel(value).time()
        except Exception:
            fraction = float(value) % 1
            seconds = int(round(fraction * 24 * 60 * 60))
            return (datetime.min + timedelta(seconds=seconds)).time()

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(cleaned, fmt).time()
            except ValueError:
                continue

    return None


def _scan_pci_factor(rows: list[tuple[Any, ...]]) -> float | None:
    for row in rows:
        for cell in row:
            if not isinstance(cell, str):
                continue
            upper = cell.upper()
            if "PCI" not in upper:
                continue
            matches = PCI_PATTERN.findall(cell)
            if matches:
                value = parse_numeric_value(matches[-1])
                if value is not None:
                    return value
    return None


def parse_scada_excel(
    file_bytes: bytes,
    *,
    interval_minutes: int = DEFAULT_INTERVAL_MINUTES,
    pci_factor_override: float | None = None,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> tuple[list[ScadaReading], ScadaParseStats]:
    if interval_minutes <= 0:
        raise ValueError("interval_minutes must be greater than zero.")

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        resolved_sheet_name = sheet_name
    else:
        worksheet = workbook.active
        resolved_sheet_name = worksheet.title

    metadata_rows = list(
        worksheet.iter_rows(min_row=1, max_row=8, values_only=True)
    )
    pci_factor = pci_factor_override or _scan_pci_factor(metadata_rows) or DEFAULT_PCI_FACTOR

    max_column = worksheet.max_column or 0
    if max_column < DATA_START_COL:
        workbook.close()
        stats = ScadaParseStats(
            total_columns=0,
            parsed_columns=0,
            skipped_columns=0,
            pci_factor=pci_factor,
            interval_minutes=interval_minutes,
            sheet_name=resolved_sheet_name,
        )
        return [], stats

    date_row = next(
        worksheet.iter_rows(
            min_row=DATE_ROW,
            max_row=DATE_ROW,
            min_col=DATA_START_COL,
            max_col=max_column,
            values_only=True,
        ),
        (),
    )
    time_row = next(
        worksheet.iter_rows(
            min_row=TIME_ROW,
            max_row=TIME_ROW,
            min_col=DATA_START_COL,
            max_col=max_column,
            values_only=True,
        ),
        (),
    )

    total_columns = max(len(date_row), len(time_row))
    if total_columns <= 0:
        workbook.close()
        stats = ScadaParseStats(
            total_columns=0,
            parsed_columns=0,
            skipped_columns=0,
            pci_factor=pci_factor,
            interval_minutes=interval_minutes,
            sheet_name=resolved_sheet_name,
        )
        return [], stats

    max_data_col = DATA_START_COL + total_columns - 1
    data_rows = list(
        worksheet.iter_rows(
            min_row=START_ROW,
            max_row=START_ROW + len(FIELD_ORDER) - 1,
            min_col=DATA_START_COL,
            max_col=max_data_col,
            values_only=True,
        )
    )
    readings: list[ScadaReading] = []
    parsed_columns = 0

    interval_hours = interval_minutes / 60

    for offset in range(total_columns):
        date_value = _parse_date(date_row[offset] if offset < len(date_row) else None)
        time_value = _parse_time(time_row[offset] if offset < len(time_row) else None)
        if date_value is None or time_value is None:
            continue

        timestamp = datetime.combine(date_value, time_value)
        metrics: dict[str, float | None] = {}
        has_value = False

        for row_offset, key in enumerate(FIELD_ORDER):
            if key is None:
                continue
            row_values = data_rows[row_offset] if row_offset < len(data_rows) else ()
            raw_value = row_values[offset] if offset < len(row_values) else None
            numeric_value = parse_numeric_value(raw_value)
            metrics[key] = numeric_value
            if numeric_value is not None:
                has_value = True

        if not has_value:
            continue

        gas_flow = metrics.get("gas_flow_nm3h")
        if gas_flow is None:
            gas_flow = 0.0
        power_gross = metrics.get("electrical_power_kw")
        if power_gross is None:
            power_gross = 0.0

        normalized_kwh = normalize_gas_to_kwh(gas_flow * interval_hours, pci_factor)
        raw_metrics = {
            key: value
            for key, value in metrics.items()
            if key not in {"gas_flow_nm3h", "electrical_power_kw"}
        }

        readings.append(
            ScadaReading(
                timestamp=timestamp,
                normalized_kwh=normalized_kwh,
                power_gross_kw=power_gross,
                gas_flow_nm3h=gas_flow,
                raw_metrics=raw_metrics,
            )
        )
        parsed_columns += 1

    workbook.close()

    stats = ScadaParseStats(
        total_columns=total_columns,
        parsed_columns=parsed_columns,
        skipped_columns=max(total_columns - parsed_columns, 0),
        pci_factor=pci_factor,
        interval_minutes=interval_minutes,
        sheet_name=resolved_sheet_name,
    )
    return readings, stats
